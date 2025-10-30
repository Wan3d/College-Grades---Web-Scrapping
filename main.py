#python -m pip install -> force installation in virtual environments
from selenium.webdriver import Chrome
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import time
from bs4 import BeautifulSoup
from credentials import usernameID, password, urlCollege, urlGrades
import xlsxwriter
import openpyxl
from email.message import EmailMessage
import ssl
import smtplib
from credentials import gmailUser, gmailPassword, gmailReceiver
from telegram_notifier import sendTelegramMessage


def initializateDriver():
    service = Service(ChromeDriverManager().install())

    option = webdriver.ChromeOptions()
    option.add_argument("--log-level=1")
    option.add_argument("--window-size=1920,1080")
    option.add_argument("--headless") # Doesn't show the window

    driver = Chrome(service = service, options = option)
    return driver

def initializateExcel():
    workbook = xlsxwriter.Workbook('data_export.xlsx')
    worksheet = workbook.add_worksheet(name="New")
    return worksheet, workbook
    
def setGrades(worksheet, listGrades):
    setHeader(worksheet)

    row = 1
    col = 0

    for i in range(len(listGrades)):
        worksheet.write(row, col, listGrades[i])
        col += 1
        if col == 13:
            row += 1
            col = 0

def setHeader(worksheet):
    listHeader = ['Materia', 'Grupo', 'I', 'II', 'III', 'IV', 'V', 'VI', 'VII', 'VIII', 'IX', 'X', 'XI']
    for col_num, header in enumerate(listHeader):
        worksheet.write(0, col_num, header)

def getExcelList():
    workbook = openpyxl.load_workbook('data_export.xlsx')
    sheet = workbook.active

    excelList = []

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for shell in row:
            excelList.append(shell.value)

    return excelList

def sendViaEmailOrTelegram(viaEmail, message):
    if viaEmail:
        # Send an alert via Email
        sendEmail(f"{message}")
    else:
        # Send an alert via Telegram bot
        sendTelegramMessage(f"{message}")

def sendEmail(isUpdated):
    # Create message
    subject = f"Grades {isUpdated}"
    body = f"""Check out now."""

    # Initiliazate object and properties
    email = EmailMessage()
    email["From"] = gmailUser
    email["To"] = gmailReceiver
    email["Subject"] = subject
    email.set_content(body)

    context = ssl.create_default_context()

    # Acces gmail server
    with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as smtp:
        smtp.login(gmailUser, gmailPassword)
        smtp.sendmail(gmailUser, gmailReceiver, email.as_string())
        print("Email sent succesfuly!")

def listToDictionary(list):
    # Creamos los datos que usaremos para nuestro diccionario
    plainList = list
    dictionaryGrades = {}
    gradesBlock = 13

    # Iteramos de 13 en 13, ya que principalmente buscaremos los nombres de la materia
    # que se repiten cada 13 índices
    for i in range(0, len(plainList), gradesBlock):
        # Del bloque de la materia, se extraen 13 columnas de esa materia en específico
        subjectBlock = plainList[i : i + 13]

        # Se identifica el nombre de la materia
        subjectKey = subjectBlock[0]

        # Agarramos el bloque de calificaciones dentro de la materia
        gradesBlock = subjectBlock[2:]

        # En un diccionario, guardamos las calificaciones asociadas al nombre de la materia
        dictionaryGrades[subjectKey] = gradesBlock
    return dictionaryGrades

def subjectName(id, list):
    for i in range(0, len(list), 13):
        word = list[i]
        if word[1:8] in id:
            list[i] = id[word[1:8]]
    return list

def main():
    driver = initializateDriver()
    driver.get(urlCollege)

    user_input = driver.find_element(By.ID, "usr")
    user_input.send_keys(usernameID) 

    password_input = driver.find_element(By.ID, "pwd")
    password_input.send_keys(password)

    button = driver.find_element(By.XPATH, "/html/body/div[2]/div/form/div[3]/div/button")
    button.click()

    secondButton = driver.find_element(By.XPATH, "/html/body/div/nav/div/div[2]/ul/li[2]/a")
    secondButton.click()

    thirdButton = driver.find_element(By.XPATH, "/html/body/div/nav/div/div[2]/ul/li[2]/ul/li[5]/a")
    thirdButton.click()

    time.sleep(5)

    soup = BeautifulSoup(driver.page_source, 'html.parser')

    # asignatures = soup.find_all('b')
    # Syntax list comprehension -> list = [expression for item in iterable if condition == True]
    # listAsignatures = [asignature.get_text().replace('\xa0', '').replace('\t', ' ') for asignature in asignatures]

    listGrades = []

    # Get all the tags tr with id non
    tr = soup.find_all('tr', id="non")

    contCols = 0

    # Iterate for every tag tr
    for row in tr:
        td = row.find_all('td') # Get td column where grades are set
        for column in td: # Ignore the first index because is already listed
            contCols += 1
            # Replace with None grades without a value
            if contCols >= 2 and contCols <= 13:
                if column.get_text() == '':
                    listGrades.append(column.get_text().replace('', 'None'))
                else:
                    listGrades.append(column.get_text().replace('\xa0', '').replace('\t', ' '))
            else:
                listGrades.append(column.get_text().replace('\xa0', '').replace('\t', ' '))
        contCols = 0

    # Get a base list from Excel with grades. This is a starting point to start comparing
    excelList = getExcelList()

    # !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!! #

    from id_subject import idSubject
    # Maps ID with subject namez
    excelList = subjectName(idSubject, excelList)
    listGrades = subjectName(idSubject, listGrades)

    # Converts both lists to dictionary type
    dicOldGrades = listToDictionary(excelList)
    dicNewGrades = listToDictionary(listGrades)
    
    # !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!! #

    if listGrades != excelList:
        # Update new grades
        worksheet, workbook = initializateExcel()
        setGrades(worksheet, listGrades)
        workbook.close()

        # Looks up for changes
        for subject, newGrade in dicNewGrades.items():
            if subject in dicOldGrades:
                oldGrade = dicOldGrades[subject]

                for j, (old, new) in enumerate(zip(oldGrade, newGrade)):
                    if old != new:
                        sendViaEmailOrTelegram(False, 
                                        f"Materia -> {subject}\nUnidad ->  {j + 1}\nCalificación -> {new}")

    driver.quit()

if __name__ == "__main__":
    main()
